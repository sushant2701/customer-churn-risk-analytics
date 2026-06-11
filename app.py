import os
import json
import pandas as pd
from io import BytesIO
from flask import Flask, render_template, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import db_manager
import etl_pipeline
import ml_model

app = Flask(__name__)

# Ensure folder structure exists
os.makedirs(app.instance_path, exist_ok=True)

@app.route('/')
@app.route('/dashboard')
def dashboard():
    """Renders the main customer churn risk dashboard."""
    # Check database status
    db_exists = os.path.exists(db_manager.DB_FILE)
    
    # Initialize default UI metrics
    metrics = {
        "total_customers": "N/A",
        "churn_rate": "N/A",
        "high_risk_count": "N/A",
        "avg_tenure": "N/A"
    }
    top_customers = []
    model_metadata = ml_model.get_model_metadata()
    model_trained = model_metadata is not None
    
    if db_exists:
        try:
            conn = db_manager.get_db_connection()
            cursor = conn.cursor()
            
            # Fetch Total Customers
            cursor.execute("SELECT COUNT(*) FROM customers;")
            total_cust = cursor.fetchone()[0]
            
            if total_cust > 0:
                metrics["total_customers"] = f"{total_cust:,}"
                
                # Fetch Churn Rate (Actual)
                cursor.execute("SELECT COUNT(*) FROM usage_metrics WHERE churn_label = 1;")
                churned_cust = cursor.fetchone()[0]
                metrics["churn_rate"] = f"{(churned_cust / total_cust) * 100:.2f}%"
                
                # Fetch Avg Tenure
                cursor.execute("SELECT AVG(tenure_months) FROM customers;")
                avg_ten = cursor.fetchone()[0]
                metrics["avg_tenure"] = f"{avg_ten:.1f} months"
                
                # Fetch High Risk Count (predicted churn probability >= 0.5)
                cursor.execute("SELECT COUNT(*) FROM usage_metrics WHERE churn_prob >= 0.5;")
                high_risk = cursor.fetchone()[0]
                metrics["high_risk_count"] = f"{high_risk:,}" if model_trained else "Train Model"
                
                # Fetch Top 20 High-Risk Customers (ordered by churn probability)
                cursor.execute("""
                    SELECT c.customer_id, c.tenure_months, c.billing_type, 
                           u.usage_gb, u.monthly_charges, u.total_charges, u.num_complaints, u.churn_label, u.churn_prob
                    FROM customers c
                    JOIN usage_metrics u ON c.customer_id = u.customer_id
                    WHERE u.churn_prob IS NOT NULL
                    ORDER BY u.churn_prob DESC
                    LIMIT 20;
                """)
                top_customers = [dict(row) for row in cursor.fetchall()]
                
            conn.close()
        except Exception as e:
            print(f"Error querying dashboard metrics: {e}")
            
    return render_template(
        'dashboard.html',
        metrics=metrics,
        top_customers=top_customers,
        model_trained=model_trained,
        db_exists=db_exists,
        model_metadata=model_metadata
    )

@app.route('/api/dashboard/charts')
def api_dashboard_charts():
    """Returns aggregated data for Chart.js dashboard charts."""
    db_exists = os.path.exists(db_manager.DB_FILE)
    if not db_exists:
        return jsonify({"error": "Database not initialized"})
        
    try:
        conn = db_manager.get_db_connection()
        
        # 1. Billing Type distribution
        billing_df = pd.read_sql_query("""
            SELECT billing_type, COUNT(*) as count 
            FROM customers 
            GROUP BY billing_type;
        """, conn)
        
        # 2. Churn rate by Usage Tier
        # Low: usage_gb < 150, Medium: 150-350, High: > 350
        usage_df = pd.read_sql_query("""
            SELECT 
                CASE 
                    WHEN u.usage_gb < 150 THEN 'Low (<150 GB)'
                    WHEN u.usage_gb BETWEEN 150 AND 350 THEN 'Medium (150-350 GB)'
                    ELSE 'High (>350 GB)'
                END as usage_tier,
                COUNT(*) as total,
                SUM(u.churn_label) as churned
            FROM usage_metrics u
            GROUP BY usage_tier;
        """, conn)
        
        conn.close()
        
        # Format usage metrics
        usage_labels = []
        usage_rates = []
        usage_counts = []
        for _, row in usage_df.iterrows():
            usage_labels.append(row['usage_tier'])
            rate = (row['churned'] / row['total']) * 100 if row['total'] > 0 else 0
            usage_rates.append(round(rate, 2))
            usage_counts.append(int(row['churned']))
            
        chart_data = {
            "billing": {
                "labels": billing_df['billing_type'].tolist(),
                "values": billing_df['count'].tolist()
            },
            "usage_churn": {
                "labels": usage_labels,
                "rates": usage_rates,
                "counts": usage_counts
            }
        }
        return jsonify(chart_data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/etl')
def etl_page():
    """Renders the ETL pipeline logging and execution panel."""
    db_exists = os.path.exists(db_manager.DB_FILE)
    stats = {
        "raw_exists": os.path.exists(os.path.join('data', 'raw_customers.csv')),
        "cleaned_exists": os.path.exists(os.path.join('data', 'cleaned_customers.csv')),
        "db_records": 0
    }
    
    if db_exists:
        try:
            conn = db_manager.get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM customers;")
            stats["db_records"] = cursor.fetchone()[0]
            conn.close()
        except Exception:
            pass
            
    return render_template('etl.html', stats=stats)

@app.route('/api/etl/logs')
def api_etl_logs():
    """Fetches text logs generated by the ETL pipeline for display in the console."""
    log_path = os.path.join('data', 'etl.log')
    if os.path.exists(log_path):
        with open(log_path, 'r') as f:
            logs = f.read()
        return jsonify({"logs": logs})
    return jsonify({"logs": "No logs recorded yet. Click 'Run ETL Pipeline' to start."})

@app.route('/api/etl/run', methods=['POST'])
def api_etl_run():
    """Triggers the raw data generation and ETL execution workflow."""
    try:
        # Step 1: Generate Raw Data
        raw_csv = etl_pipeline.generate_raw_data(50000)
        
        # Step 2: Run ETL Pipeline (Clean and Normalize)
        df_cleaned, stats = etl_pipeline.run_etl(raw_csv)
        
        # Step 3: Initialize Database Schema and Load data
        db_manager.init_db()
        cust_cnt, metrics_cnt = db_manager.load_data_to_db(df_cleaned)
        
        # Step 4: If a model was already trained, we remove metadata to prompt retraining
        metadata_file = os.path.join('instance', 'model_metadata.json')
        if os.path.exists(metadata_file):
            os.remove(metadata_file)
        model_file = os.path.join('instance', 'random_forest_model.pkl')
        if os.path.exists(model_file):
            os.remove(model_file)
            
        stats['db_customers'] = cust_cnt
        stats['db_metrics'] = metrics_cnt
        
        return jsonify({"success": True, "stats": stats})
    except Exception as e:
        etl_pipeline.log_message(f"ERROR during ETL run: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/sql')
def sql_page():
    """Handles query segmentation with filter parameters, pagination, and SQL explains."""
    billing_type = request.args.get('billing_type', 'All')
    usage_tier = request.args.get('usage_tier', 'All')
    page = request.args.get('page', 1, type=int)
    per_page = 15
    
    db_exists = os.path.exists(db_manager.DB_FILE)
    rows = []
    total_records = 0
    explain_plan = []
    sql_query = ""
    total_pages = 0
    
    if db_exists:
        try:
            rows, total_records, explain_plan, sql_query = db_manager.get_segmented_customers(
                billing_type, usage_tier, page=page, per_page=per_page
            )
            total_pages = (total_records + per_page - 1) // per_page
        except Exception as e:
            print(f"Error executing SQL search: {e}")
            
    return render_template(
        'sql_segments.html',
        db_exists=db_exists,
        rows=rows,
        billing_type=billing_type,
        usage_tier=usage_tier,
        page=page,
        total_records=total_records,
        total_pages=total_pages,
        explain_plan=explain_plan,
        sql_query=sql_query
    )

@app.route('/model')
def model_page():
    """Displays ML Model dashboard with performance indicators and feature importances."""
    metadata = ml_model.get_model_metadata()
    db_exists = os.path.exists(db_manager.DB_FILE)
    return render_template('model.html', metadata=metadata, db_exists=db_exists)

@app.route('/api/model/train', methods=['POST'])
def api_model_train():
    """Triggers Scikit-learn Random Forest model training and database back-scoring."""
    try:
        metadata = ml_model.train_churn_model()
        if "error" in metadata:
            return jsonify({"success": False, "error": metadata["error"]}), 400
        return jsonify({"success": True, "metadata": metadata})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/model/predict', methods=['POST'])
def api_model_predict():
    """Scores a single user manual input for custom churn prediction testing."""
    try:
        data = request.json
        tenure = float(data.get('tenure_months', 0))
        monthly = float(data.get('monthly_charges', 0))
        total = float(data.get('total_charges', 0))
        complaints = int(data.get('num_complaints', 0))
        usage = float(data.get('usage_gb', 0))
        billing_type = data.get('billing_type', 'Month-to-month')
        
        result = ml_model.predict_single_customer(
            tenure=tenure,
            monthly=monthly,
            total=total,
            complaints=complaints,
            usage=usage,
            billing_type=billing_type
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/export')
def export_high_risk():
    """Downloads high risk customers list (probability >= 0.5) as a styled Excel spreadsheet."""
    db_exists = os.path.exists(db_manager.DB_FILE)
    if not db_exists:
        return "Database not initialized. Run ETL first.", 400
        
    try:
        conn = db_manager.get_db_connection()
        query = """
        SELECT c.customer_id, c.tenure_months, c.billing_type,
               u.usage_gb, u.monthly_charges, u.total_charges, u.num_complaints, u.churn_label, u.churn_prob
        FROM customers c
        JOIN usage_metrics u ON c.customer_id = u.customer_id
        WHERE u.churn_prob >= 0.5
        ORDER BY u.churn_prob DESC;
        """
        df = pd.read_sql_query(query, conn)
        conn.close()
        
        if len(df) == 0:
            return "No high-churn-risk records found (make sure model is trained).", 400
            
        # Create spreadsheet workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "High Churn Risk Report"
        
        # Add headers
        headers = [
            "Customer ID", "Tenure (Months)", "Billing Type", 
            "Monthly Usage (GB)", "Monthly Charges", "Total Charges", 
            "Complaints Count", "Actual Churn Status", "Churn Risk Probability"
        ]
        ws.append(headers)
        
        # Header Styling
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2A4B7C", end_color="2A4B7C", fill_type="solid") # Executive Navy
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Set professional column widths (fixed)
        col_widths = {
            'A': 15,  # Customer ID
            'B': 16,  # Tenure
            'C': 18,  # Billing Type
            'D': 20,  # Monthly Usage
            'E': 18,  # Monthly Charges
            'F': 18,  # Total Charges
            'G': 18,  # Complaints Count
            'H': 20,  # Actual Churn Status
            'I': 22   # Churn Risk Probability
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
            
        # Populate sheet with pre-formatted strings for speed
        for index, row in df.iterrows():
            ws.append([
                row['customer_id'],
                int(row['tenure_months']),
                row['billing_type'],
                f"{row['usage_gb']:.1f} GB",
                f"${row['monthly_charges']:.2f}",
                f"${row['total_charges']:.2f}",
                int(row['num_complaints']),
                "Churned" if int(row['churn_label']) == 1 else "Active",
                f"{row['churn_prob']*100:.2f}%"
            ])
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="telecom_high_churn_risk_list.xlsx"
        )
    except Exception as e:
        return f"Error during report compilation: {str(e)}", 500

if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5050)
